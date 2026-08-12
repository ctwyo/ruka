"""
Palm recognition – kiosk-mode server.

Camera runs on the SERVER, frames are streamed to all browsers as MJPEG.
WebSocket pushes scan/registration state to all connected clients.

Open from any device on the LAN:  http://<server-ip>:8000

Run:  python server.py
"""

import asyncio
import contextlib
import io
import json
import os
import logging
from typing import List, Optional
from pydantic import BaseModel, Field
import socket
import time
from datetime import datetime, timezone, date as date_cls
from pathlib import Path
import cv2
import numpy as np

import mediapipe as mp
from mediapipe.tasks import python
from mediapipe.tasks.python import vision
# from mediapipe.framework.formats import landmark_pb2
# from mediapipe.python.solutions import drawing_utils as mp_drawing
# from mediapipe.python.solutions import drawing_styles as mp_drawing_styles
# Замена старого drawing_utils и landmark_pb2
from mediapipe.tasks.python.vision import drawing_utils as mp_drawing
from mediapipe.tasks.python.vision import drawing_styles as mp_drawing_styles
from mediapipe.tasks.python.vision import HandLandmarker, HandLandmarksConnections

import torch
import uvicorn
from PIL import Image, ImageDraw, ImageFont
from fastapi import FastAPI, WebSocket, WebSocketDisconnect, HTTPException
from fastapi.responses import HTMLResponse, StreamingResponse, Response

from compnet.model import compnet, preprocess_palm

os.environ["GLOG_minloglevel"] = "3"
os.environ["TF_CPP_MIN_LOG_LEVEL"] = "3"

# ── constants ────────────────────────────────────────────────────────────────
EXTRACTION_FILE_PATH = Path(__file__).parent / "import.txt"
DB_PATH = Path(__file__).parent / "palms.json"
ATTENDANCE_PATH = Path(__file__).parent / "attendance.json"
WEIGHTS_PATH = Path(__file__).parent / "compnet" / "weights.pth"
SAMPLES = 150               # кадров на регистрацию (~10 сек при 15 fps)
TEMPLATES_PER_USER = 5      # 5 шаблонов из 150 кадров = 1 шаблон каждые ~2 сек
# LOCK_THRESHOLD = 0.75     # старое значение
LOCK_THRESHOLD = 0.95      # порог косинусной близости: ниже — не свой 87
# MARGIN = 0.05             # старое значение
MARGIN = 0.08               # минимальный отрыв первого кандидата от второго
MIN_SCAN_FRAMES = 5         # минимум кадров перед локом (защита от случайного матча за 1 кадр)
FRAMES_PER_ATTEMPT = 15     # как часто инкрементится номер «попытки» в UI (~1 сек при 15 fps)
HAND_LOST_FRAMES = 6        # сколько кадров без руки до сброса сканирования
ROI_PAD = 1.3               # размер кропа относительно длины ладони (wrist→middle MCP)
HULL_EXPAND = 1.18          # на сколько раздуть выпуклую оболочку ладони наружу от центра
MASK_DILATE_FRAC = 0.04     # дилатация маски в долях от размера кропа
EMBED_DIM = 512             # размерность embedding-вектора CompNet
CAMERA_INDEX = 0            # индекс веб-камеры (0 = первая)
JPEG_QUALITY = 80           # качество JPEG для MJPEG-стрима (1–100)
TARGET_FPS = 15             # целевой FPS обработки кадров
TARGET_RECT_RATIO = 0.75    # сторона квадрата-прицела = ratio × высоты кадра

# ── MediaPipe ────────────────────────────────────────────────────────────────
# mp_hands = mp.solutions.hands
# mp_draw = mp.solutions.drawing_utils
# mp_styles = mp.solutions.drawing_styles
# _hands = mp_hands.Hands(
#     model_complexity=1,
#     min_detection_confidence=0.7,
#     min_tracking_confidence=0.5,
#     max_num_hands=1,
#     static_image_mode=False,
# )

# 1. Задаем путь к скачанному файлу модели
base_options = python.BaseOptions(model_asset_path='hand_landmarker.task')

# 2. Настраиваем конфигурацию (полный аналог ваших старых параметров)
options = vision.HandLandmarkerOptions(
    base_options=base_options,
    running_mode=vision.RunningMode.VIDEO,  # False в static_image_mode означает режим видео/камеры
    min_hand_detection_confidence=0.7,     # min_detection_confidence
    min_hand_presence_confidence=0.5,      # min_tracking_confidence (для детекции)
    min_tracking_confidence=0.5,           # min_tracking_confidence (для трекинга)
    num_hands=1                            # max_num_hands
)

# 3. Создаем объект детектора
_hands = vision.HandLandmarker.create_from_options(options)

# ── CNN: CompNet (palm-print specialist) ─────────────────────────────────────
print(f"Loading CompNet weights from {WEIGHTS_PATH.name}...")
_cnn = compnet(num_classes=600)   # 600 = Tongji training classes; not used at inference
_state = torch.load(str(WEIGHTS_PATH), map_location="cpu", weights_only=True)
_cnn.load_state_dict(_state, strict=True)
_cnn.eval()
print("CompNet ready.")

# Настройка логирования
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger("api")


# ── ROI + embedding ──────────────────────────────────────────────────────────
def extract_palm_roi(frame: np.ndarray, landmarks):
    """Returns (crop_bgr, palm_mask_uint8) or (None, None).

    palm_mask is 255 inside the convex hull of palm landmarks (slightly
    expanded), 0 outside — so downstream code can drop the background.
    """
    h, w = frame.shape[:2]
    pts = np.array([[lm.x * w, lm.y * h] for lm in landmarks], dtype=np.float32)
    wrist, middle_mcp = pts[0], pts[9]
    palm_size = float(np.linalg.norm(middle_mcp - wrist))
    if palm_size < 20:
        return None, None
    center = (wrist + middle_mcp) * 0.5
    dx, dy = middle_mcp - wrist
    angle = float(np.degrees(np.arctan2(dy, dx))) + 90
    side = int(palm_size * ROI_PAD)
    M = cv2.getRotationMatrix2D(tuple(center), angle, 1.0)
    rotated = cv2.warpAffine(frame, M, (w, h), flags=cv2.INTER_LINEAR,
                             borderMode=cv2.BORDER_REPLICATE)

    # Apply the same rotation to all 21 landmarks so we can build a mask in
    # crop-space coordinates.
    pts_h = np.hstack([pts, np.ones((len(pts), 1), dtype=np.float32)])  # (21, 3)
    rotated_pts = (M @ pts_h.T).T                                       # (21, 2)

    half = side // 2
    x1, y1 = int(center[0]) - half, int(center[1]) - half
    x2, y2 = x1 + side, y1 + side
    pad_l, pad_t = max(0, -x1), max(0, -y1)
    pad_r, pad_b = max(0, x2 - w), max(0, y2 - h)
    x1c, y1c = max(0, x1), max(0, y1)
    x2c, y2c = min(w, x2), min(h, y2)
    crop = rotated[y1c:y2c, x1c:x2c]
    if crop.size == 0:
        return None, None
    if any((pad_l, pad_r, pad_t, pad_b)):
        crop = cv2.copyMakeBorder(crop, pad_t, pad_b, pad_l, pad_r,
                                  cv2.BORDER_REPLICATE)

    # Palm hull: wrist + thumb base + finger MCPs. Excludes fingers/background.
    # 0=wrist, 1=thumb CMC, 2=thumb MCP, 5/9/13/17=index/middle/ring/pinky MCP.
    palm_idx = [0, 1, 2, 5, 9, 13, 17]
    crop_pts = rotated_pts[palm_idx] - np.array([x1, y1], dtype=np.float32)
    hull_center = crop_pts.mean(axis=0)
    expanded = hull_center + (crop_pts - hull_center) * HULL_EXPAND
    hull = cv2.convexHull(expanded.astype(np.int32))

    ch, cw = crop.shape[:2]
    mask = np.zeros((ch, cw), dtype=np.uint8)
    cv2.fillConvexPoly(mask, hull, 255)
    # Soft margin: dilation covers palm flesh that extends past landmark joints.
    k = max(3, int(round(min(ch, cw) * MASK_DILATE_FRAC)))
    mask = cv2.dilate(mask, np.ones((k, k), np.uint8))
    return crop, mask


def compute_embedding(roi: np.ndarray, mask: np.ndarray | None = None,
                      debug_dir: str | None = None) -> np.ndarray | None:
    if roi is None or roi.size == 0:
        return None
    x = preprocess_palm(roi, mask=mask, size=128, debug_dir=debug_dir)  # (1, 1, 128, 128)
    with torch.no_grad():
        feat = _cnn.getFeatureCode(x).squeeze(0).numpy()
    # getFeatureCode already L2-normalizes
    return feat.astype(np.float32)


def cosine(a: np.ndarray, b: np.ndarray) -> float:
    return float(np.dot(a, b))


def _entry_templates(entry: dict) -> list[np.ndarray]:
    raw = entry.get("feats") or [entry["feat"]]
    return [np.array(f, dtype=np.float32) for f in raw]


# ── vectorized template cache ────────────────────────────────────────────────
# Stack ALL templates (5 per user × N users) into one big (T, 512) matrix.
# Matching becomes a single BLAS call:  scores = matrix @ feat
# Per-user max is computed via np.maximum.at on indices.
def _rebuild_template_cache() -> None:
    db = state["db"]
    cache = state["tpl_cache"]
    if not db:
        cache["matrix"] = None
        cache["user_idx"] = None
        cache["names"] = []
        return

    names: list[str] = []
    rows: list[list[float]] = []
    indices: list[int] = []
    for u_idx, (name, entry) in enumerate(db.items()):
        names.append(name)
        raw = entry.get("feats") or [entry["feat"]]
        for vec in raw:
            rows.append(vec)
            indices.append(u_idx)

    cache["matrix"] = np.asarray(rows, dtype=np.float32)        # (T, 512)
    cache["user_idx"] = np.asarray(indices, dtype=np.int32)     # (T,)
    cache["names"] = names                                       # length n_users
    print(f"[cache] rebuilt: {len(names)} users, {cache['matrix'].shape[0]} templates")


def _vectorized_per_user_max(feat: np.ndarray) -> tuple[list[str], np.ndarray] | None:
    """Returns (names, max_score_per_user). None if DB empty."""
    cache = state["tpl_cache"]
    if cache["matrix"] is None:
        return None
    scores = cache["matrix"] @ feat                             # (T,) BLAS
    n_users = len(cache["names"])
    per_user_max = np.full(n_users, -1.0, dtype=np.float32)
    np.maximum.at(per_user_max, cache["user_idx"], scores)
    return cache["names"], per_user_max


def scan_update(feat: np.ndarray | None, db: dict) -> dict | None:
    sc = state["scan"]
    if feat is None:
        sc["missing"] += 1
        if sc["missing"] >= HAND_LOST_FRAMES:
            _reset_scan()
        return sc["locked"]

    sc["missing"] = 0
    sc["seen_frames"] += 1
    if not db:
        return None

    res = _vectorized_per_user_max(feat)          # ← BLAS-vectorized
    if res is None:
        return None
    names, scores = res

    # initialise / re-init peaks if user count changed
    if sc["peaks_arr"] is None or len(sc["peaks_arr"]) != len(names):
        sc["peaks_arr"] = np.full(len(names), -1.0, dtype=np.float32)
        sc["peaks_names"] = names
    # vectorized "running max"
    np.maximum(sc["peaks_arr"], scores, out=sc["peaks_arr"])

    # top-2 via argpartition (O(n) — no full sort even at 50K users)
    arr = sc["peaks_arr"]
    if len(arr) == 1:
        best_idx, second_peak = 0, -1.0
    else:
        # negate so we get the largest two
        idx2 = np.argpartition(-arr, 1)[:2]
        if arr[idx2[0]] >= arr[idx2[1]]:
            best_idx, second_idx = int(idx2[0]), int(idx2[1])
        else:
            best_idx, second_idx = int(idx2[1]), int(idx2[0])
        second_peak = float(arr[second_idx])
    best_name = sc["peaks_names"][best_idx]
    best_peak = float(arr[best_idx])
    margin_ok = second_peak < 0 or (best_peak - second_peak) >= MARGIN

    if sc["locked"] is not None:
        return sc["locked"]

    if sc["seen_frames"] >= MIN_SCAN_FRAMES and best_peak >= LOCK_THRESHOLD and margin_ok:
        sc["locked"] = {
            "name": best_name, "score": round(best_peak, 4),
            "second": round(second_peak, 4) if second_peak > -1 else None,
            "ok": True, "scanning": False,
        }
        return sc["locked"]

    return {
        "name": best_name, "score": round(best_peak, 4),
        "second": round(second_peak, 4) if second_peak > -1 else None,
        "ok": False, "scanning": True,
        "attempt": sc["seen_frames"] // FRAMES_PER_ATTEMPT + 1,
    }


# ── db ───────────────────────────────────────────────────────────────────────
def load_db() -> dict:
    if not DB_PATH.exists():
        return {}
    try:
        data = json.loads(DB_PATH.read_text(encoding="utf-8"))
    except Exception as e:
        print(f"DB load error: {e}; starting empty")
        return {}
    if not data:
        return {}
    first = next(iter(data.values()))
    # legacy: list values were 210-d distance vectors
    if isinstance(first, list):
        DB_PATH.write_text("{}", encoding="utf-8")
        print("[!] Old DB format (distance vectors) -> cleared.")
        return {}
    # legacy: MobileNet-era 960-d / 1280-d embeddings — incompatible with CompNet
    feats = first.get("feats") or [first.get("feat")]
    if feats and len(feats[0]) != EMBED_DIM:
        DB_PATH.write_text("{}", encoding="utf-8")
        print(f"[!] Old DB ({len(feats[0])}-d embeddings) -> cleared. "
              f"Re-register palms with CompNet ({EMBED_DIM}-d).")
        return {}
    return data


def save_db(db: dict) -> None:
    DB_PATH.write_text(json.dumps(db, indent=2), encoding="utf-8")


def _users_list(db: dict) -> list[dict]:
    return [{"name": k} for k in db.keys()]


# ── attendance ───────────────────────────────────────────────────────────────
def load_attendance() -> dict:
    if not ATTENDANCE_PATH.exists():
        return {}
    try:
        return json.loads(ATTENDANCE_PATH.read_text(encoding="utf-8"))
    except Exception:
        return {}


def save_attendance(att: dict) -> None:
    ATTENDANCE_PATH.write_text(
        json.dumps(att, indent=2, ensure_ascii=False), encoding="utf-8"
    )


def record_event(name: str, kind: str) -> dict:
    """kind ∈ {'in','out'}. 'in' is set only if missing (first scan wins);
    'out' overwrites (latest scan wins).
    Returns the actually-stored time + a 'changed' flag for the UI."""
    assert kind in ("in", "out")
    now = datetime.now()
    day = now.strftime("%Y-%m-%d")
    t = now.strftime("%H:%M:%S")

    att = state["attendance"]
    day_entries = att.setdefault(day, {})
    entry = day_entries.setdefault(name, {"in": None, "out": None})

    # TEST MODE: both in and out overwrite on every scan
    if kind == "in":
        entry["in"] = t
    else:
        entry["out"] = t

    save_attendance(att)
    return {"day": day, "name": name, "kind": kind,
            "time": t, "changed": True, "entry": entry}


# ── shared state ─────────────────────────────────────────────────────────────
state = {
    "db": load_db(),
    "attendance": load_attendance(),
    "tpl_cache": {"matrix": None, "user_idx": None, "names": []},
    "reg": {"active": False, "name": None, "buffer": []},
    "scan": {"peaks_arr": None, "peaks_names": None,
             "seen_frames": 0, "missing": 0, "locked": None, "acted": False},
    # attendance intent: None | "in" | "out". When set, the next confident lock
    # records that event for the recognized user and clears the intent.
    "intent": None,
    # last-event toast (broadcast to clients, shown briefly in the UI)
    "last_event": None,         # {name, kind, time, day, ts (monotonic)}
    "latest_jpeg": None,
    "clients": set(),
    "camera_ok": False,
    "save_debug": False,
    "last_debug_dir": None,
}


def _reset_scan():
    state["scan"].update(peaks_arr=None, peaks_names=None,
                         seen_frames=0, missing=0, locked=None, acted=False)


# Build cache once on startup
_rebuild_template_cache()


# ── per-frame processing (runs in executor thread) ───────────────────────────
def handle_frame(frame: np.ndarray) -> tuple[bytes | None, dict]:
    frame = cv2.flip(frame, 1)  # mirror for selfie view

    out: dict = {
        "hand": False, "match": None, "register": None,
        "users": _users_list(state["db"]),
    }

    # 1. Конвертируем кадр OpenCV (BGR) в формат MediaPipe Image
    mp_image = mp.Image(image_format=mp.ImageFormat.SRGB, data=frame)

    # 2. Получаем временную метку кадра в миллисекундах (критично для режима VIDEO)
    # Если у вас есть объект cap (cv2.VideoCapture), лучше использовать: int(cap.get(cv2.CAP_PROP_POS_MSEC))
    # В качестве универсального решения используем системное время:
    # frame_timestamp_ms = int(datetime.utcnow().timestamp() * 1000)
    frame_timestamp_ms = int(datetime.now(timezone.utc).timestamp() * 1000)

    # 3. Запускаем распознавание через новый метод
    result = _hands.detect_for_video(mp_image, frame_timestamp_ms)

    # debug dump request (consumed once)
    debug_dir = None
    if state["save_debug"]:
        debug_dir = str(Path(__file__).parent / "debug" /
                        datetime.now().strftime("%Y%m%d_%H%M%S"))
        os.makedirs(debug_dir, exist_ok=True)
        # save the un-annotated frame as stage 0 (before any landmarks drawn)
        cv2.imwrite(os.path.join(debug_dir, "0_full_frame.png"), frame)
        state["save_debug"] = False
        state["last_debug_dir"] = debug_dir

    feat = None
    
    # В новом API поле называется 'hand_landmarks' вместо 'multi_hand_landmarks'
    # if result.hand_landmarks:
    #     # ИСПРАВЛЕНО: Берем именно первую найденную руку (индекс [0])
    #     # Теперь hand — это список объектов NormalizedLandmark
    #     hand = result.hand_landmarks[0]  
    #     h_, w_ = frame.shape[:2]
        
    #     # Теперь lm.x и lm.y отработают корректно без ошибок
    #     pts = np.array([[lm.x * w_, lm.y * h_] for lm in hand],
    #                    dtype=np.float32)
    #     out["palm_bbox"] = [float(pts[:, 0].min()), float(pts[:, 1].min()),
    #                         float(pts[:, 0].max()), float(pts[:, 1].max())]
        
    #     # Передаем точки в ваши кастомные функции обработки ладони
    #     roi, palm_mask = extract_palm_roi(frame, hand)
    #     feat = compute_embedding(roi, mask=palm_mask, debug_dir=debug_dir)
        
    #     # 4. Отрисовка точек через новые утилиты Tasks API
    #     mp_drawing.draw_landmarks(
    #         image=frame,                         
    #         landmark_list=hand,                  
    #         connections=HandLandmarksConnections.HAND_CONNECTIONS, 
    #         landmark_drawing_spec=mp_drawing_styles.get_default_hand_landmarks_style(),
    #         connection_drawing_spec=mp_drawing_styles.get_default_hand_connections_style(),
    #                     # Передаем размеры кадра (ширину и высоту) для не-квадратных видеопотоков:
    #         image_dimensions=ImageDimensions(width=w_, height=h_)
    #     )
    #     out["hand"] = True



    # if result.hand_landmarks:
    #         # 1. Объявите индексы связей скелета ладони (за пределами или внутри условия)
    #     HAND_CONNECTIONS = [
    #         (0, 1), (1, 2), (2, 3), (3, 4),       # Большой палец
    #         (0, 5), (5, 6), (6, 7), (7, 8),       # Указательный
    #         (5, 9), (9, 10), (10, 11), (11, 12),  # Средний
    #         (9, 13), (13, 14), (14, 15), (15, 16),# Безымянный
    #         (13, 17), (0, 17), (17, 18), (18, 19), (19, 20) # Мизинец и ладонь
    #     ]
    #     # result.hand_landmarks[0] — список из 21 точки первой руки
    #     hand = result.hand_landmarks[0]  
    #     h_, w_ = frame.shape[:2]
        
    #     # Расчет bbox и эмбеддингов (ваш оригинальный код)
    #     pts = np.array([[lm.x * w_, lm.y * h_] for lm in hand], dtype=np.float32)
    #     out["palm_bbox"] = [float(pts[:, 0].min()), float(pts[:, 1].min()),
    #                         float(pts[:, 0].max()), float(pts[:, 1].max())]
        
    #     roi, palm_mask = extract_palm_roi(frame, hand)
    #     feat = compute_embedding(roi, mask=palm_mask, debug_dir=debug_dir)
        
    #     # 2. ИСПРАВЛЕНИЕ: Отрисовка на чистом OpenCV вместо mp_drawing.draw_landmarks
    #     # Переводим нормализованные координаты всех 21 точек в физические пиксели кадра
    #     pixel_pts = [(int(lm.x * w_), int(lm.y * h_)) for lm in hand]
        
    #     # Рисуем соединительные линии (зеленый цвет, толщина 2)
    #     for connection in HAND_CONNECTIONS:
    #         start_idx, end_idx = connection
    #         cv2.line(frame, pixel_pts[start_idx], pixel_pts[end_idx], (0, 255, 0), 2)
            
    #     # Рисуем суставы/точки поверх линий (красный цвет, радиус 4)
    #     for pt in pixel_pts:
    #         cv2.circle(frame, pt, 4, (0, 0, 255), -1)

    #     out["hand"] = True
    if result.hand_landmarks:
        # Словарь, где для каждого пальца задан список его связей и цвет в формате BGR
        FINGERS_DATA = {
            "thumb": {  # Большой палец
                "connections": [(0, 1), (1, 2), (2, 3), (3, 4)],
                "color": (255, 0, 255)  # Пурпурный
            },
            "index": {  # Указательный
                "connections": [(0, 5), (5, 6), (6, 7), (7, 8)],
                "color": (255, 0, 0)    # Синий
            },
            "middle": { # Средний
                "connections": [(5, 9), (9, 10), (10, 11), (11, 12)],
                "color": (0, 255, 0)    # Зеленый
            },
            "ring": {   # Безымянный
                "connections": [(9, 13), (13, 14), (14, 15), (15, 16)],
                "color": (0, 255, 255)  # Желтый
            },
            "pinky": {  # Мизинец и основание ладони
                "connections": [(13, 17), (0, 17), (17, 18), (18, 19), (19, 20)],
                "color": (0, 165, 255)  # Оранжевый
            }
        }

        # В новом API Tasks берем список точек для первой руки
        hand = result.hand_landmarks[0] 
        h_, w_ = frame.shape[:2]
        
        # Расчет bbox и эмбеддингов
        pts = np.array([[lm.x * w_, lm.y * h_] for lm in hand], dtype=np.float32)
        out["palm_bbox"] = [float(pts[:, 0].min()), float(pts[:, 1].min()),
                            float(pts[:, 0].max()), float(pts[:, 1].max())]
        
        roi, palm_mask = extract_palm_roi(frame, hand)
        feat = compute_embedding(roi, mask=palm_mask, debug_dir=debug_dir)
        
        # Переводим нормализованные координаты всех 21 точек в физические пиксели кадра
        pixel_pts = [(int(lm.x * w_), int(lm.y * h_)) for lm in hand]
        
        # 1. Сначала рисуем линии для каждого пальца своим цветом
        for finger_name, data in FINGERS_DATA.items():
            finger_color = data["color"]
            for connection in data["connections"]:
                start_idx, end_idx = connection
                cv2.line(frame, pixel_pts[start_idx], pixel_pts[end_idx], finger_color, 1)
                
        # 2. Затем рисуем точки (суставы). Чтобы они соответствовали цвету пальца,
        # мы красим их в зависимости от их индекса (ID от 0 до 20)
        for idx, pt in enumerate(pixel_pts):
            # По умолчанию для запястья (ID 0) используем белый цвет
            pt_color = (255, 255, 255) 
            
            # # Определяем, к какому пальцу относится текущая точка
            # if idx in [1, 2, 3, 4]:
            #     pt_color = FINGERS_DATA["thumb"]["color"]
            # elif idx in [5, 6, 7, 8]:
            #     pt_color = FINGERS_DATA["index"]["color"]
            # elif idx in [9, 10, 11, 12]:
            #     pt_color = FINGERS_DATA["middle"]["color"]
            # elif idx in [13, 14, 15, 16]:
            #     pt_color = FINGERS_DATA["ring"]["color"]
            # elif idx in [17, 18, 19, 20]:
            #     pt_color = FINGERS_DATA["pinky"]["color"]
                
            # Рисуем саму точку
            cv2.circle(frame, pt, 4, pt_color, -1)

        out["hand"] = True





    reg = state["reg"]
    db = state["db"]

    if reg["active"]:
        if feat is not None:
            reg["buffer"].append(feat)
        progress = len(reg["buffer"])
        done = progress >= SAMPLES
        out["register"] = {
            "name": reg["name"], "progress": progress,
            "total": SAMPLES, "done": done,
        }
        if done:
            buf = np.stack(reg["buffer"], axis=0)
            step = max(1, len(buf) // TEMPLATES_PER_USER)
            picks = buf[::step][:TEMPLATES_PER_USER]
            picks = picks / (np.linalg.norm(picks, axis=1, keepdims=True) + 1e-9)
            db[reg["name"]] = {"feats": picks.astype(np.float32).tolist()}
            save_db(db)
            _rebuild_template_cache()
            out["users"] = _users_list(db)
            reg.update(active=False, name=None, buffer=[])
            _reset_scan()
    elif db:
        out["match"] = scan_update(feat, db)

        # ─── intent commit: record check-in/out the FIRST time a lock fires ──
        m = out["match"]
        sc = state["scan"]
        if (m and m.get("ok") and state["intent"] is not None
                and not sc.get("acted")):
            ev = record_event(m["name"], state["intent"])
            sc["acted"] = True
            state["intent"] = None
            state["last_event"] = {**ev, "ts": time.monotonic()}

    out["intent"] = state["intent"]
    # show event toast for 4 seconds after recording
    le = state["last_event"]
    if le and (time.monotonic() - le["ts"]) < 4.0:
        out["last_event"] = {k: v for k, v in le.items() if k != "ts"}
    out["attendance_today"] = state["attendance"].get(
        datetime.now().strftime("%Y-%m-%d"), {}
    )

    _draw_overlay(frame, out)

    ok, jpeg = cv2.imencode(".jpg", frame, [cv2.IMWRITE_JPEG_QUALITY, JPEG_QUALITY])
    return (jpeg.tobytes() if ok else None), out


# ── PIL-based text rendering (supports Cyrillic) ────────────────────────────
_FONT_CACHE: dict[int, ImageFont.FreeTypeFont] = {}


def _font(size: int) -> ImageFont.FreeTypeFont:
    if size in _FONT_CACHE:
        return _FONT_CACHE[size]
    for name in ("arial.ttf", "segoeui.ttf",
                 "C:\\Windows\\Fonts\\arial.ttf",
                 "C:\\Windows\\Fonts\\segoeui.ttf"):
        try:
            f = ImageFont.truetype(name, size)
            _FONT_CACHE[size] = f
            return f
        except OSError:
            continue
    _FONT_CACHE[size] = ImageFont.load_default()
    return _FONT_CACHE[size]


def _draw_overlay(frame: np.ndarray, out: dict) -> None:
    """All overlays go through PIL so Cyrillic renders correctly."""
    h, w = frame.shape[:2]
    pil = Image.fromarray(cv2.cvtColor(frame, cv2.COLOR_BGR2RGB))
    draw = ImageDraw.Draw(pil, "RGBA")

    def text_size(s: str, fnt) -> tuple[int, int]:
        bbox = draw.textbbox((0, 0), s, font=fnt)
        return bbox[2] - bbox[0], bbox[3] - bbox[1]

    # ─── target rectangle (corner brackets) ──
    side = int(h * TARGET_RECT_RATIO)
    rx0 = (w - side) // 2
    ry0 = (h - side) // 2
    rx1, ry1 = rx0 + side, ry0 + side

    palm_bbox = out.get("palm_bbox")
    rect_color = (120, 200, 255)        # default — empty
    rect_hint = "Поднесите ладонь в рамку"
    if palm_bbox:
        cx = (palm_bbox[0] + palm_bbox[2]) / 2
        cy = (palm_bbox[1] + palm_bbox[3]) / 2
        palm_w = palm_bbox[2] - palm_bbox[0]
        palm_h = palm_bbox[3] - palm_bbox[1]
        in_rect = rx0 <= cx <= rx1 and ry0 <= cy <= ry1
        # rough size check: palm should fill at least ~40% of rect
        size_ok = max(palm_w, palm_h) >= side * 0.4
        if in_rect and size_ok:
            rect_color = (80, 255, 130)
            rect_hint = None
        elif in_rect and not size_ok:
            rect_color = (255, 200, 60)
            rect_hint = "Ближе к камере"
        else:
            rect_color = (255, 200, 60)
            rect_hint = "Сдвиньте ладонь в центр"

    corner = 36
    thickness = 4
    for (x, y, dx1, dy1, dx2, dy2) in [
        (rx0, ry0,  corner,  0,       0,       corner),     # top-left
        (rx1, ry0, -corner,  0,       0,       corner),     # top-right
        (rx0, ry1,  corner,  0,       0,      -corner),     # bot-left
        (rx1, ry1, -corner,  0,       0,      -corner),     # bot-right
    ]:
        draw.line([(x, y), (x + dx1, y + dy1)], fill=rect_color, width=thickness)
        draw.line([(x, y), (x + dx2, y + dy2)], fill=rect_color, width=thickness)

    if rect_hint:
        fnt = _font(18)
        tw, th = text_size(rect_hint, fnt)
        tx, ty = (w - tw) // 2, ry1 + 8
        draw.rectangle([tx - 8, ty - 4, tx + tw + 8, ty + th + 4], fill=(0, 0, 0))
        draw.text((tx, ty), rect_hint, fill=rect_color, font=fnt)

    def banner(text: str, color_rgb: tuple[int, int, int]):
        fnt = _font(22)
        tw, th = text_size(text, fnt)
        draw.rectangle([0, 0, w, th + 20], fill=(0, 0, 0))
        draw.text(((w - tw) // 2, 8), text, fill=color_rgb, font=fnt)

    def toast(text: str, color_rgb: tuple[int, int, int]):
        fnt = _font(26)
        tw, th = text_size(text, fnt)
        x0, y0 = (w - tw) // 2 - 14, h // 2 - th // 2 - 10
        x1, y1 = (w + tw) // 2 + 14, h // 2 + th // 2 + 10
        draw.rectangle([x0, y0, x1, y1], fill=(0, 0, 0, 220))
        draw.text(((w - tw) // 2, h // 2 - th // 2), text, fill=color_rgb, font=fnt)

    def bottom(text: str, color_rgb: tuple[int, int, int], sub: str | None = None):
        fnt = _font(20)
        sub_fnt = _font(14)
        tw, th = text_size(text, fnt)
        sw, sh = (0, 0) if not sub else text_size(sub, sub_fnt)
        pad = 8
        box_w = max(tw, sw) + 2 * pad
        y_bottom = h - 10
        y_top = y_bottom - th - (sh + 4 if sub else 0) - 2 * pad
        draw.rectangle([4, y_top, 4 + box_w, y_bottom], fill=(0, 0, 0))
        if sub:
            draw.text((4 + pad, y_top + pad), sub, fill=(200, 200, 200), font=sub_fnt)
            draw.text((4 + pad, y_top + pad + sh + 4), text, fill=color_rgb, font=fnt)
        else:
            draw.text((4 + pad, y_top + pad), text, fill=color_rgb, font=fnt)

    # ─── intent banner ──
    intent = out.get("intent")
    if intent == "in":
        banner("ПРИХОД — поднесите ладонь", (100, 255, 0))
    elif intent == "out":
        banner("УХОД — поднесите ладонь", (255, 90, 90))

    # ─── recording toast ──
    le = out.get("last_event")
    if le:
        kind_label = "приход" if le["kind"] == "in" else "уход"
        if le.get("changed", True):
            msg = f"{le['name']} — {kind_label}  {le['time']}"
            color_rgb = (120, 255, 120)
        else:
            msg = f"{le['name']} — {kind_label} уже был в {le['time']}"
            color_rgb = (255, 200, 60)
        toast(msg, color_rgb)

    # ─── bottom status ──
    r = out["register"]
    m = out["match"]
    if r is not None:
        if r["done"]:
            bottom(f"Сохранён: {r['name']}", (100, 255, 100))
        else:
            bottom(f"Запись {r['name']}: {r['progress']}/{r['total']}",
                   (255, 200, 60),
                   sub="Медленно подвигайте рукой ближе/дальше, чуть наклоняйте")
    elif m is not None:
        score_pct = m["score"] * 100
        if m.get("scanning"):
            attempt = m.get("attempt", 1)
            bottom(f"Сканирую #{attempt}…  {score_pct:.0f}%", (255, 200, 60))
        elif m["ok"]:
            bottom(f"MATCH: {m['name']}  {score_pct:.0f}%", (100, 255, 100))
        else:
            bottom(f"Неизвестно  {score_pct:.0f}%", (255, 100, 100))
    elif not out["hand"]:
        bottom("Нет руки", (140, 140, 140))
    else:
        bottom("Рука найдена — БД пуста", (180, 180, 180))

    # write back into the frame buffer
    frame[:] = cv2.cvtColor(np.array(pil), cv2.COLOR_RGB2BGR)


# ── camera loop ──────────────────────────────────────────────────────────────
def _try_open_camera(index: int) -> cv2.VideoCapture | None:
    """Try DSHOW first (Windows-friendly), then default backend."""
    for backend in (cv2.CAP_DSHOW, cv2.CAP_ANY):
        cap = cv2.VideoCapture(index, backend)
        if cap.isOpened():
            cap.set(cv2.CAP_PROP_FRAME_WIDTH, 640)
            cap.set(cv2.CAP_PROP_FRAME_HEIGHT, 480)
            ok, _ = cap.read()
            if ok:
                return cap
            cap.release()
    return None


async def camera_loop():
    loop = asyncio.get_event_loop()

    # try CAMERA_INDEX first, then 0..3 as fallback
    candidates = [CAMERA_INDEX] + [i for i in range(4) if i != CAMERA_INDEX]
    cap = None
    for idx in candidates:
        cap = await loop.run_in_executor(None, _try_open_camera, idx)
        if cap is not None:
            state["camera_ok"] = True
            print(f"Camera #{idx} opened.")
            break

    if cap is None:
        print("[!] Cannot open any camera. Possible causes:")
        print("    • another app is using it (Zoom, Teams, browser tab with getUserMedia)")
        print("    • Windows camera privacy disabled for desktop apps")
        print("    • no camera connected")
        return

    period = 1.0 / TARGET_FPS
    frame_count = 0
    try:
        while True:
            t0 = loop.time()
            try:
                ok, frame = await loop.run_in_executor(None, cap.read)
                if not ok or frame is None:
                    await asyncio.sleep(0.05)
                    continue

                jpeg, status = await loop.run_in_executor(None, handle_frame, frame)
                if jpeg is not None:
                    state["latest_jpeg"] = jpeg
                    frame_count += 1
                    if frame_count in (1, 10, 100):
                        print(f"[camera] produced frame #{frame_count}, "
                              f"jpeg={len(jpeg)} bytes, clients={len(state['clients'])}")

                for ws in list(state["clients"]):
                    try:
                        await ws.send_text(json.dumps(status))
                    except Exception:
                        state["clients"].discard(ws)
            except Exception as e:
                import traceback
                print(f"[camera] frame error: {e}")
                traceback.print_exc()
                await asyncio.sleep(0.1)

            dt = loop.time() - t0
            if dt < period:
                await asyncio.sleep(period - dt)
    finally:
        cap.release()
        print("[camera] released")



# Схема JSON-ответа (пользователь)
class UserResponse(BaseModel):
    code: int = Field(serialization_alias="code")
    name: str = Field(serialization_alias="name")
    profile_code: int = Field(serialization_alias="profile-code")
    position_fio: Optional[str] = Field(default=None, serialization_alias="position-fio")
    password: Optional[str] = Field(default=None, serialization_alias="password")
    user_card: Optional[str] = Field(default=None, serialization_alias="user-card")
    inn: Optional[str] = Field(default=None, serialization_alias="inn")


# Класс парсинга (читает файл с диска в кодировке Windows-1251)
class Extract:
    def __init__(self):
        self.users: List[UserResponse] = []

    def process_file(self, file_path: str):
        logger.info(f"extraction.ProcessFile: {file_path}")
        
        if not os.path.exists(file_path):
            raise FileNotFoundError(f"File not found: {file_path}")

        with open(file_path, "r", encoding="cp1251") as f:
            content = f.read()

        ss = [line.rstrip("\r") for line in content.split("\n")]
        if not ss or ss == [""]:
            raise ValueError("empty extraction")

        for i, uu in enumerate(ss):
            if uu == "$$$DELETEALLUSERS":
                self.users = []
            elif uu == "$$$ADDUSERS":
                self.add_users(ss[i + 1 :])
            elif uu == "$$$DELETEUSERSBYCODE":
                self.del_user_by_code(ss[i + 1 :])

    def del_user_by_code(self, ss: List[str]):
        for code in ss:
            if code.startswith("$$$"):
                return
            if not code.strip():
                continue
            try:
                cc = int(code)
            except ValueError as e:
                raise ValueError(f"wrong field Code({code}) in DelUser: {e}")
            self.users = [u for u in self.users if u.code != cc]

    def add_users(self, ss: List[str]):
        for i, uu in enumerate(ss):
            if not uu.strip():
                continue
            
            data = uu.split(";")
            if len(data) > 1:
                if data[0].startswith("$$$"):
                    return
                try:
                    code = int(data[0])
                except ValueError as e:
                    raise ValueError(f"wrong field Code({data[0]}) in User[{i+1}]: {e}")
            else:
                if data[0].startswith("$$$") or data[0] == "":
                    return
                continue

            try:
                p_code = int(data[3]) if len(data) > 3 and data[3] else 0
            except ValueError as e:
                raise ValueError(f"wrong field Profile Code({data[3]}) in User[{i+1}]: {e}")

            user = UserResponse(
                code=code,
                name=data[1] if len(data) > 1 else "",
                profile_code=p_code,
                position_fio=data[2] if len(data) > 2 and data[2] else None,
                password=data[4] if len(data) > 4 and data[4] else None,
                user_card=data[5] if len(data) > 5 and data[5] else None,
                inn=data[6] if len(data) > 6 and data[6] else None
            )
            self.users.append(user)


# ── FastAPI ──────────────────────────────────────────────────────────────────
@contextlib.asynccontextmanager
async def lifespan(_app):
    task = asyncio.create_task(camera_loop())
    yield
    task.cancel()


app = FastAPI(lifespan=lifespan)
STATIC = Path(__file__).parent / "static"
STATIC.mkdir(exist_ok=True)


@app.get("/")
async def index():
    return HTMLResponse((STATIC / "index.html").read_text(encoding="utf-8"))

# Эндпоинт FastAPI (принимает пустой POST-запрос и парсит файл из константы)
@app.get("/parse-extraction", response_model=List[UserResponse], response_model_exclude_none=True)
async def parse_extraction():
    try:
        extractor = Extract()
        # Вызов метода парсинга с использованием константы
        extractor.process_file(EXTRACTION_FILE_PATH)
        
        # FastAPI автоматически превратит список моделей в JSON,
        # применит алиасы (profile-code) и скроет пустые поля (exclude_none=True)
        return extractor.users

    except FileNotFoundError as fnf:
        logger.error(f"File error: {fnf}")
        raise HTTPException(status_code=404, detail=str(fnf))
    except ValueError as ve:
        logger.error(f"Validation error: {ve}")
        raise HTTPException(status_code=400, detail=str(ve))
    except Exception as e:
        logger.error(f"Server error: {e}")
        raise HTTPException(status_code=500, detail=f"Internal server error: {str(e)}")



@app.get("/api/users")
async def get_users():
    return {"users": _users_list(state["db"])}


@app.delete("/api/users/{name}")
async def delete_user(name: str):
    state["db"].pop(name, None)
    save_db(state["db"])
    _rebuild_template_cache()
    _reset_scan()
    return {"users": _users_list(state["db"])}


@app.post("/api/register/{name}")
async def start_register(name: str):
    state["reg"].update(active=True, name=name, buffer=[])
    return {"status": "started", "name": name}


@app.post("/api/register/cancel")
async def cancel_register():
    state["reg"].update(active=False, name=None, buffer=[])
    return {"status": "cancelled"}


@app.post("/api/debug/save")
async def debug_save():
    """Dump the next processed frame's pipeline stages to debug/<timestamp>/."""
    state["save_debug"] = True
    return {"status": "queued"}


# ─── attendance ──────────────────────────────────────────────────────────────
@app.post("/api/attendance/mode/{kind}")
async def set_intent(kind: str):
    if kind not in ("in", "out"):
        return {"error": "kind must be 'in' or 'out'"}
    state["intent"] = kind
    _reset_scan()           # restart scan so the next palm registers fresh
    return {"intent": kind}


@app.post("/api/attendance/cancel")
async def cancel_intent():
    state["intent"] = None
    return {"intent": None}


@app.get("/api/attendance/today")
async def attendance_today():
    day = datetime.now().strftime("%Y-%m-%d")
    return {"day": day, "entries": state["attendance"].get(day, {})}


@app.get("/api/attendance/all")
async def attendance_all():
    return state["attendance"]


def _hms_to_min(s: str | None) -> int | None:
    if not s:
        return None
    try:
        h, m, sec = s.split(":")
        return int(h) * 60 + int(m) + int(sec) // 60
    except Exception:
        return None


def _build_xlsx() -> bytes:
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Attendance"

    headers = ["Дата", "Пользователь", "Приход", "Уход", "Длительность"]
    ws.append(headers)
    head_font = Font(bold=True, color="FFFFFF")
    head_fill = PatternFill("solid", fgColor="1E1E2E")
    for col_idx in range(1, len(headers) + 1):
        c = ws.cell(row=1, column=col_idx)
        c.font = head_font
        c.fill = head_fill
        c.alignment = Alignment(horizontal="center")

    rows = []
    for day, entries in sorted(state["attendance"].items()):
        for name, entry in entries.items():
            t_in = entry.get("in")
            t_out = entry.get("out")
            duration = ""
            if t_in and t_out:
                m_in = _hms_to_min(t_in)
                m_out = _hms_to_min(t_out)
                if m_in is not None and m_out is not None and m_out > m_in:
                    d = m_out - m_in
                    duration = f"{d // 60}ч {d % 60:02d}м"
            rows.append([day, name, t_in or "—", t_out or "—", duration])

    for r in rows:
        ws.append(r)

    # column widths
    widths = [12, 20, 12, 12, 14]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[chr(64 + i)].width = w

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


@app.get("/api/attendance/export.xlsx")
async def export_xlsx():
    data = _build_xlsx()
    fname = f"attendance_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    return Response(
        content=data,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )


@app.get("/stream")
async def stream():
    async def gen():
        boundary = b"--frame\r\n"
        while True:
            jpeg = state.get("latest_jpeg")
            if jpeg:
                yield (boundary
                       + b"Content-Type: image/jpeg\r\n"
                       + b"Content-Length: " + str(len(jpeg)).encode() + b"\r\n\r\n"
                       + jpeg + b"\r\n")
            await asyncio.sleep(1.0 / TARGET_FPS)
    return StreamingResponse(gen(), media_type="multipart/x-mixed-replace; boundary=frame")


@app.websocket("/ws")
async def ws_endpoint(websocket: WebSocket):
    await websocket.accept()
    state["clients"].add(websocket)
    try:
        # initial snapshot
        await websocket.send_text(json.dumps({
            "hand": False, "match": None, "register": None,
            "users": _users_list(state["db"]),
        }))
        while True:
            await websocket.receive_text()  # ignore pings/messages
    except WebSocketDisconnect:
        pass
    finally:
        state["clients"].discard(websocket)


def _list_ipv4() -> list[str]:
    """List all IPv4 addresses on this machine, skipping loopback.
    Falls back to `ipconfig` on Windows if getaddrinfo returns too few."""
    ips: set[str] = set()
    try:
        for info in socket.getaddrinfo(socket.gethostname(), None, socket.AF_INET):
            ip = info[4][0]
            if not ip.startswith("127."):
                ips.add(ip)
    except Exception:
        pass
    if os.name == "nt":
        try:
            import re, subprocess
            out = subprocess.check_output(["ipconfig"], text=True,
                                          encoding="utf-8", errors="ignore", timeout=3)
            ips.update(re.findall(r"IPv4.*?:\s*(\d+\.\d+\.\d+\.\d+)", out))
        except Exception:
            pass
    ips.discard("127.0.0.1")
    return sorted(ips)


if __name__ == "__main__":
    print("\n  Local:  http://localhost:8000")
    addrs = _list_ipv4()
    if addrs:
        print("  LAN candidates (pick the one matching your Wi-Fi/Ethernet subnet):")
        for ip in addrs:
            print(f"          http://{ip}:8000")
    print()
    uvicorn.run(app, host="0.0.0.0", port=8000, log_level="warning")
